import os
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from gestion.models import Persona, Alumno, Docente, Carrera, Materia, PlanEstudio, Comision, ComisionDocente, Cursada, Evaluacion

class Command(BaseCommand):
    help = 'Poblar la base de datos a partir de los archivos Excel en la carpeta Recursos'

    @transaction.atomic
    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS("Iniciando sembrado de datos desde Excel..."))
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        recursos_dir = os.path.join(base_dir, 'Recursos')
        if not os.path.exists(recursos_dir):
            self.stdout.write(self.style.WARNING("Carpeta 'Recursos' no encontrada. Omitiendo sembrado de Excel en este entorno."))
            return

        carreras_file = os.path.join(recursos_dir, 'Carreras_y_Materias_ISFT188.xlsx')
        if os.path.exists(carreras_file):
            self.stdout.write("Procesando Carreras_y_Materias_ISFT188.xlsx...")
            wb = openpyxl.load_workbook(carreras_file, data_only=True)
            
            # Sheet: Carreras_ISFT188
            if 'Carreras_ISFT188' in wb.sheetnames:
                ws_carreras = wb['Carreras_ISFT188']
                for r in range(5, ws_carreras.max_row + 1):
                    codigo = ws_carreras.cell(r, 1).value
                    nombre = ws_carreras.cell(r, 2).value
                    resolucion = ws_carreras.cell(r, 3).value
                    if codigo and nombre:
                        Carrera.objects.update_or_create(
                            codigo_carrera=str(codigo).strip(),
                            defaults={
                                'nombre_carrera': str(nombre).strip(),
                                'resolucion_vigente': str(resolucion).strip() if resolucion else ''
                            }
                        )

            # Sheet: Materias_Planes_Estudio
            if 'Materias_Planes_Estudio' in wb.sheetnames:
                ws_materias = wb['Materias_Planes_Estudio']
                for r in range(4, ws_materias.max_row + 1):
                    carrera_nombre = ws_materias.cell(r, 1).value
                    resolucion = ws_materias.cell(r, 2).value
                    cod_mat = ws_materias.cell(r, 3).value
                    anio = ws_materias.cell(r, 4).value
                    nombre_mat = ws_materias.cell(r, 5).value
                    carga_anual = ws_materias.cell(r, 6).value
                    carga_semanal = ws_materias.cell(r, 7).value
                    correlativas = ws_materias.cell(r, 8).value
                    docente_nombre = ws_materias.cell(r, 9).value

                    if carrera_nombre and cod_mat and nombre_mat:
                        c_norm = str(carrera_nombre).strip()
                        carrera_obj = Carrera.objects.filter(nombre_carrera__icontains=c_norm[:12]).first()
                        if not carrera_obj:
                            codigo_gen = c_norm.upper()[:12].replace(" ", "_")
                            carrera_obj, _ = Carrera.objects.get_or_create(
                                codigo_carrera=codigo_gen,
                                defaults={'nombre_carrera': c_norm, 'resolucion_vigente': str(resolucion or '')}
                            )

                        cod_mat_str = str(cod_mat).strip()
                        materia_obj, _ = Materia.objects.get_or_create(
                            codigo_materia=f"{carrera_obj.codigo_carrera}_{cod_mat_str}",
                            defaults={'nombre_materia': str(nombre_mat).strip()}
                        )

                        try:
                            anio_int = int(anio)
                        except (ValueError, TypeError):
                            anio_int = 1

                        try:
                            c_anual = int(carga_anual) if carga_anual else None
                        except (ValueError, TypeError):
                            c_anual = None

                        try:
                            c_semanal = int(carga_semanal) if carga_semanal else None
                        except (ValueError, TypeError):
                            c_semanal = None

                        plan_obj, _ = PlanEstudio.objects.get_or_create(
                            carrera=carrera_obj,
                            materia=materia_obj,
                            defaults={
                                'anio_carrera': anio_int,
                                'carga_horaria_anual': c_anual,
                                'carga_horaria_semanal': c_semanal,
                                'correlatividades': str(correlativas or '').strip()
                            }
                        )

                        # Docente asignado al plan
                        if docente_nombre and str(docente_nombre).strip() not in ['- - - - - - - - - -', 'None', '']:
                            doc_clean = str(docente_nombre).strip()
                            parts = doc_clean.split(" ")
                            apellido_doc = parts[0]
                            nombre_doc = " ".join(parts[1:]) if len(parts) > 1 else "Docente"
                            dni_doc = f"DOC_{abs(hash(doc_clean)) % 1000000}"
                            
                            p_doc, _ = Persona.objects.get_or_create(
                                dni=dni_doc,
                                defaults={'nombre': nombre_doc, 'apellido': apellido_doc}
                            )
                            d_doc, _ = Docente.objects.get_or_create(persona=p_doc)
                            
                            com, _ = Comision.objects.get_or_create(
                                codigo_comision=f"COM_{plan_obj.id_plan}_2026",
                                defaults={'plan_estudio': plan_obj, 'anio_lectivo': 2026}
                            )
                            ComisionDocente.objects.get_or_create(comision=com, docente=d_doc, defaults={'rol': 'Titular'})

        # 2. Cargar Planillas de Alumnos de Logística y Prácticas Deportivas
        archivos_alumnos = [
            ('13 - Tec. Y Sist. De Inf. En Logistica 3.xlsx', 'LOGISTICA'),
            ('15 - Tics en las Pract. Deportivas.xlsx', 'DEPORTIVAS')
        ]

        for file_name, tag in archivos_alumnos:
            file_path = os.path.join(recursos_dir, file_name)
            if not os.path.exists(file_path):
                continue
            self.stdout.write(f"Procesando planilla de alumnos: {file_name}...")
            wb = openpyxl.load_workbook(file_path, data_only=True)

            for sheet_name in wb.sheetnames:
                if 'Avance' not in sheet_name:
                    continue
                ws = wb[sheet_name]
                
                # Extraer datos de la cabecera del acta
                carrera_raw = str(ws.cell(7, 1).value or ws.cell(7, 2).value or "").replace("Carrera:", "").strip()
                materia_raw = str(ws.cell(8, 2).value or ws.cell(8, 1).value or "").replace("Espacio curricular:", "").strip()
                docente_raw = str(ws.cell(9, 2).value or ws.cell(9, 1).value or "").replace("Docente a cargo:", "").strip()

                if not materia_raw or materia_raw == 'None':
                    materia_raw = file_name.replace(".xlsx", "")

                carrera_obj = Carrera.objects.filter(nombre_carrera__icontains="Logística" if tag == 'LOGISTICA' else "Deportivas").first()
                if not carrera_obj:
                    carrera_obj = Carrera.objects.first()

                materia_obj, _ = Materia.objects.get_or_create(
                    codigo_materia=f"MAT_{tag}_{abs(hash(materia_raw)) % 100000}",
                    defaults={'nombre_materia': materia_raw}
                )

                plan_obj, _ = PlanEstudio.objects.get_or_create(
                    carrera=carrera_obj,
                    materia=materia_obj,
                    defaults={'anio_carrera': 3 if '3' in file_name else 1}
                )

                comision_obj, _ = Comision.objects.get_or_create(
                    codigo_comision=f"COM_{tag}_{plan_obj.id_plan}_2025",
                    defaults={'plan_estudio': plan_obj, 'anio_lectivo': 2025}
                )

                if docente_raw and docente_raw != 'None':
                    parts = docente_raw.split(" ")
                    ap_d = parts[0]
                    nom_d = " ".join(parts[1:]) if len(parts) > 1 else "Docente"
                    p_doc, _ = Persona.objects.get_or_create(
                        dni=f"DOC_{abs(hash(docente_raw)) % 1000000}",
                        defaults={'nombre': nom_d, 'apellido': ap_d}
                    )
                    doc_obj, _ = Docente.objects.get_or_create(persona=p_doc)
                    ComisionDocente.objects.get_or_create(comision=comision_obj, docente=doc_obj)

                # Iterar filas de alumnos (de fila 13 en adelante)
                for r in range(13, ws.max_row + 1):
                    dni_val = ws.cell(r, 3).value
                    ap_nom = ws.cell(r, 4).value

                    if not dni_val or not ap_nom or str(ap_nom).strip() == '':
                        continue

                    # Normalizar DNI
                    try:
                        dni_clean = str(int(float(str(dni_val).strip())))
                    except (ValueError, TypeError):
                        continue

                    ap_nom_str = str(ap_nom).strip()
                    if ',' in ap_nom_str:
                        parts = ap_nom_str.split(',')
                        apellido = parts[0].strip()
                        nombre = parts[1].strip()
                    else:
                        parts = ap_nom_str.split(' ')
                        apellido = parts[0].strip()
                        nombre = " ".join(parts[1:]).strip() if len(parts) > 1 else "Alumno"

                    # Crear o actualizar Persona y Alumno
                    persona_obj, created = Persona.objects.get_or_create(
                        dni=dni_clean,
                        defaults={
                            'nombre': nombre,
                            'apellido': apellido,
                            'mail': f"{nombre.lower().replace(' ', '')}.{apellido.lower().replace(' ', '')}@alumnos.isft188.edu.ar"
                        }
                    )
                    if not created and persona_obj.nombre != nombre:
                        persona_obj.nombre = nombre
                        persona_obj.apellido = apellido
                        persona_obj.save()

                    alumno_obj, _ = Alumno.objects.get_or_create(
                        persona=persona_obj,
                        defaults={'legajo': f"LEG-{dni_clean}"}
                    )

                    # Asistencia %
                    asist_val = ws.cell(r, 14).value or ws.cell(r, 7).value
                    try:
                        asistencia = float(asist_val) * 100 if asist_val and float(asist_val) <= 1.0 else (float(asist_val) if asist_val else 85.0)
                    except (ValueError, TypeError):
                        asistencia = 80.0

                    # Situación Final
                    promociona = ws.cell(r, 5).value
                    a_final = ws.cell(r, 6).value
                    recursa = ws.cell(r, 7).value

                    situacion = 'Regular'
                    if promociona and str(promociona).strip() == '1':
                        situacion = 'Promocionado'
                    elif a_final and str(a_final).strip() == '1':
                        situacion = 'Final'
                    elif recursa and str(recursa).strip() == '1':
                        situacion = 'Libre'

                    cursada_obj, _ = Cursada.objects.get_or_create(
                        comision=comision_obj,
                        alumno=alumno_obj,
                        defaults={
                            'porcentaje_asistencia': asistencia,
                            'situacion_final': situacion
                        }
                    )

                    # Nota final
                    nota_final = ws.cell(r, 12).value
                    if nota_final is not None:
                        try:
                            nota_num = float(nota_final)
                            Evaluacion.objects.get_or_create(
                                cursada=cursada_obj,
                                instancia='Nota Final',
                                defaults={'nota': nota_num}
                            )
                        except (ValueError, TypeError):
                            pass

        total_personas = Persona.objects.count()
        total_alumnos = Alumno.objects.count()
        total_carreras = Carrera.objects.count()
        total_materias = Materia.objects.count()
        total_cursadas = Cursada.objects.count()

        self.stdout.write(self.style.SUCCESS(f"¡Sembrado completado con éxito!"))
        self.stdout.write(f"Resumen DB:")
        self.stdout.write(f"- Total Personas: {total_personas}")
        self.stdout.write(f"- Total Alumnos: {total_alumnos}")
        self.stdout.write(f"- Total Carreras: {total_carreras}")
        self.stdout.write(f"- Total Materias: {total_materias}")
        self.stdout.write(f"- Total Cursadas: {total_cursadas}")
